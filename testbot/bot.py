#!/usr/bin/env python
import os
try:import openpyxl
except ImportError:os.system('python -m pip install openpyxl')
try:import pandas
except ImportError:os.system('python -m pip install pandas')
import openpyxl
from ast import Global
from flask import Flask,render_template,request
import threading
import time
import requests
from openpyxl import Workbook
import datetime


in_l1=0 ;in_u1=1;in_l2=2;in_u2=3;in_l3=4;in_u3=5
inplace=[0,0,0,0,0,0,0]
TotalRS=0
l1=0.0000
l2=0.0000
l3=0.0000
u1=0.0000
u2=0.0000
u3=0.0000
tokenpair="usdtinr"
level_1_rs=TotalRS/2
level_2_rs=TotalRS/3
level_3_rs=TotalRS/5
level_1_token=0
level_2_token=0
level_3_token=0
total_tokens=0


def thread_function(name):
    global TotalRS
    global l1
    global l2
    global l3
    global u1
    global u2
    global u3
    global tokenpair
    global level_1_rs
    global level_2_rs
    global level_3_rs
    global level_1_token
    global level_2_token
    global level_3_token
    global total_tokens
    global in_l1,in_u1,in_l2,in_u2,in_l3,in_u3,inplace
    sourceFileDir = os.path.dirname(os.path.abspath(__file__))
    xlPath = os.path.join(sourceFileDir, 'out/bot_data.xlsx')


    

    book = Workbook()
    sheet = book.active
    sheet['A1'] = "Time"
    sheet['B1'] = "Value"
    sheet['C1'] = "L1"
    sheet['D1'] = "U1"
    sheet['E1'] = "l2"
    sheet['F1'] = "u2"
    sheet['G1'] = "l3"
    sheet['H1'] = "u3"
    sheet['I1'] = "level_1_rs"
    sheet['K1'] = "level_2_rs"
    sheet['M1'] = "level_3_rs"
    sheet['J1'] = "level_1_token"
    sheet['L1'] = "level_2_token"
    sheet['N1'] = "level_3_token"
    sheet['P1'] = "RS"
    sheet['O1']="Tokens"
    book.save(xlPath)
    prevtokenvalue=0
    itr=1
    while 1:
        
        data = requests.get('https://api.wazirx.com/api/v2/tickers')
        tokenvalue=float(data.json()[tokenpair]['last'])
        if prevtokenvalue==tokenvalue :
            continue
        prevtokenvalue=tokenvalue
        itr+=1
        current_datetime = datetime.datetime.now()
        
        
        
        print("Token value= ", tokenvalue)
        
        if(TotalRS<=0):
            print("0 Balance ")
            time.sleep(2)
            #continue
        if ((l1<=0)|(l2<=0)|(l3<=0)|(u1<=0)|(u2<=0)|(u3<=0)):
            print("Sell and Buy points not set")
            time.sleep(2)
            #continue

        if (tokenvalue <=l1 )and (inplace[in_l1]==0):
            print("Triggered Buy 1")
            inplace[in_l1]=1
            inplace[in_u1]=0
            level_1_token+=level_1_rs/tokenvalue
            level_1_rs=0
            print ("Bought "+str(level_1_token)+" Token") 

        if (tokenvalue <=l2 )and (inplace[in_l2]==0):
            inplace[in_l2]=1
            inplace[in_u2]=0
            level_2_token+=level_2_rs/tokenvalue
            level_2_rs=0
            print ("Bought "+str(level_2_token)+" Token") 
        if (tokenvalue <=l3 )and (inplace[in_l3]==0):
            inplace[in_l3]=1
            inplace[in_u3]=0
            level_3_token+=level_3_rs/tokenvalue
            level_3_rs=0
            print ("Bought "+str(level_1_token)+" Token") 
        if (tokenvalue >=u1 )and (inplace[in_u1]==0):
            inplace[in_u1]=1 
            inplace[in_l1]=0           
            level_1_rs+=level_1_token*tokenvalue
            level_1_token=0
            print ("Sold "+str(level_1_token)+" Token " ) 
        if (tokenvalue >=u2 )and (inplace[in_u2]==0):
            inplace[in_u2]=1 
            inplace[in_l2]=0             
            level_2_rs+=level_2_token*tokenvalue
            level_2_token=0
            print ("Sold "+str(level_2_token)+" Token " ) 
        if (tokenvalue >=u3 )and (inplace[in_u3]==0):
            inplace[in_u3]=1    
            inplace[in_l3]=0          
            level_3_rs+=level_3_token*tokenvalue
            level_3_token=0
            print ("Sold "+str(level_3_token)+" Token " ) 

        TotalRS=level_1_rs+level_2_rs+level_3_rs
        total_tokens=level_1_token+level_2_token+level_3_token
        sheet.cell(row=itr, column=1).value = current_datetime.strftime('%x %X')
        sheet.cell(row=itr, column=2).value = tokenvalue
        sheet.cell(row=itr, column=3).value = l1
        sheet.cell(row=itr, column=4).value = u1
        sheet.cell(row=itr, column=5).value = l2
        sheet.cell(row=itr, column=6).value = u2
        sheet.cell(row=itr, column=7).value = l3
        sheet.cell(row=itr, column=8).value = u3
        sheet.cell(row=itr, column=9).value = level_1_rs
        sheet.cell(row=itr, column=10).value = level_1_token
        sheet.cell(row=itr, column=11).value = level_2_rs
        sheet.cell(row=itr, column=11).value = level_2_token
        sheet.cell(row=itr, column=12).value = level_3_rs
        sheet.cell(row=itr, column=13).value = level_3_token
        sheet.cell(row=itr, column=15).value = TotalRS
        sheet.cell(row=itr, column=16).value = total_tokens
        book.save(xlPath)
        time.sleep(1)
        



app = Flask(__name__)
 
@app.route('/set')
def form():
    return render_template('set.html')
 
@app.route('/data/', methods = ['POST', 'GET'])
def data():
    global TotalRS
    global l1,level_1_rs
    global l2,level_2_rs
    global l3,level_3_rs
    global u1
    global u2
    global u3
 


    if request.method == 'GET':
        return f"The URL /data is accessed directly. Try going to '/form' to submit form"
    if request.method == 'POST':
        data = request.form

        TotalRS+=float(data['Rs'])#adding RS to existing
        print(TotalRS)
        l1=float(data['l1'])
        u1=float(data['u1'])
        l2=float(data['l2'])
        u2=float(data['u2'])
        l3=float(data['l3'])
        u3=float(data['u3'])
        level_1_rs+=TotalRS/3
        level_2_rs+=TotalRS/3
        level_3_rs+=TotalRS/3
        


        return render_template('data.html',form_data = data)
 
x = threading.Thread(target=thread_function, args=(1,))
x.start()

app.run(host='localhost', port=5000)
print("End")
